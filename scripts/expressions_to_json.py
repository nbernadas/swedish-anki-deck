"""Convert a reviewed workbook of Swedish expressions into JSON entries.

The input workbook needs two sheets: ``Partikelverb`` and ``Idioms``. Starting
at row 3, the first four columns must be Swedish expression, English meaning,
Swedish example, and English example translation. The script performs a dry
run by default and refuses to overwrite JSON files.
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


SHEETS = {
    "Partikelverb": ("partikelverb", "att", ["partikelverb", "verb"]),
    "Idioms": ("idioms", "", ["idiom", "expression"]),
}


def filename_stem(value: str) -> str:
    """Create a Windows-safe, stable filename while retaining Swedish letters."""
    value = unicodedata.normalize("NFC", value).strip().lower()
    return re.sub(r'[<>:"/\\\\|?*]+', "_", value)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True, help="Reviewed .xlsx workbook.")
    parser.add_argument("--output", type=Path, required=True, help="Folder for the generated JSON entries.")
    parser.add_argument("--write", action="store_true", help="Create files; otherwise only validate and count them.")
    args = parser.parse_args()

    if not args.input.is_file():
        parser.error(f"Workbook not found: {args.input}")
    workbook = load_workbook(args.input, read_only=True, data_only=True)

    entries: list[tuple[Path, dict[str, object]]] = []
    for sheet_name, (folder, article, tags) in SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            parser.error(f"Expected worksheet missing: {sheet_name}")
        for row_number, row in enumerate(workbook[sheet_name].iter_rows(min_row=3, values_only=True), start=3):
            swedish, definition, example, translation = (str(value or "").strip() for value in row[:4])
            if not swedish:
                continue
            if not all((definition, example, translation)):
                parser.error(f"{sheet_name} row {row_number} is incomplete: {swedish!r}")
            entry = {
                "word": swedish,
                "article": article,
                "definitions": [{
                    "definition": definition,
                    "example": example,
                    "example_translation": translation,
                }],
                "inflections": {},
                "tags": tags,
                "source_sheet": sheet_name,
                "source_order": row_number - 2,
            }
            entries.append((args.output / folder / f"{filename_stem(swedish)}.json", entry))

    collisions = [path for path, _ in entries if path.exists()]
    print(f"Prepared {len(entries)} entries; {len(collisions)} destination files already exist.")
    if not args.write:
        print("Dry run only. Re-run with --write to create the JSON files.")
        return
    if collisions:
        parser.error("Refusing to overwrite existing files. Choose an empty output folder.")
    for path, entry in entries:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(entry, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Created {len(entries)} JSON files under {args.output}.")


if __name__ == "__main__":
    main()
